# Copyright 2026 Private AI Ltd (SC881079)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Deterministic artefact validators. Run after
generation so a corrupt artefact never ships to a client: a failure raises
`ValidationError` (a RuntimeError → ML 503) rather than returning a broken file.

DOCX: the package is a well-formed zip carrying the required OOXML parts, and it
reopens via python-docx (which parses `document.xml`). PDF: pypdf opens it and it
has at least one page. Full ECMA XSD validation against bundled schemas is the
Phase-3 hardening; the reopen + opens-clean checks here catch the corruption that
actually occurs (truncated writes, broken zips, unparsable XML)."""

from __future__ import annotations

import json
import logging
import re
import zipfile
from pathlib import Path

log = logging.getLogger(__name__)


class ValidationError(RuntimeError):
    """Raised when a generated artefact fails its deterministic checks."""


def validate_docx(path: str) -> None:
    p = Path(path)
    if not p.is_file() or p.stat().st_size == 0:
        raise ValidationError("DOCX is missing or empty")
    try:
        with zipfile.ZipFile(p) as z:
            names = set(z.namelist())
            corrupt = z.testzip()
    except zipfile.BadZipFile as e:
        raise ValidationError(f"DOCX is not a valid OOXML package: {e}") from e
    if corrupt is not None:
        raise ValidationError(f"DOCX zip is corrupt at {corrupt}")
    for required in ("[Content_Types].xml", "word/document.xml"):
        if required not in names:
            raise ValidationError(f"DOCX is missing required part {required}")
    try:
        import docx

        docx.Document(str(p))  # parses document.xml — catches malformed body XML
    except Exception as e:
        raise ValidationError(f"DOCX failed to reopen: {e}") from e

    # ECMA/ISO XSD validation. Honest-degrade: a schema violation is logged,
    # not raised — the reopen check above is the hard gate, and we do not reject a
    # document a real Word processor would accept. (Tests call validate_docx_xsd
    # directly to assert real output passes.)
    from .config import settings

    if settings.docx_xsd_validate:
        try:
            validate_docx_xsd(path)
        except ValidationError as e:
            log.warning("DOCX XSD validation: %s", e)


# --- ECMA/ISO OOXML XSD validation -------------------------------------------

_SCHEMA_DIR = Path(__file__).parent / "assets" / "ooxml-schemas"
_WML_XSD = _SCHEMA_DIR / "ISO-IEC29500-4_2016" / "wml.xsd"

# The OOXML 2006 namespaces the WML schema knows. Everything else in a real
# document.xml — markup-compatibility (`mc:`) and the w14/w15/… extension namespaces
# it marks Ignorable — is stripped before validation (those are, by design, not in
# the schema). These URIs are the published standard; the cleaning code is our own.
_ALLOWED_OOXML_NS = frozenset({
    "http://schemas.openxmlformats.org/officeDocument/2006/math",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "http://schemas.openxmlformats.org/officeDocument/2006/sharedTypes",
    "http://schemas.openxmlformats.org/schemaLibrary/2006/main",
    "http://schemas.openxmlformats.org/drawingml/2006/main",
    "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "http://schemas.openxmlformats.org/drawingml/2006/chartDrawing",
    "http://schemas.openxmlformats.org/drawingml/2006/diagram",
    "http://schemas.openxmlformats.org/drawingml/2006/picture",
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "http://schemas.openxmlformats.org/presentationml/2006/main",
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "http://www.w3.org/XML/1998/namespace",
})

_wml_schema = None  # compiled lazily (the 39-file schema set is slow to compile)


def _ns_of(qname: str) -> str | None:
    return qname.split("}", 1)[0][1:] if isinstance(qname, str) and qname.startswith("{") else None


def _strip_ignorable(root) -> None:
    """Remove markup-compatibility / extension-namespace attributes + elements that
    the transitional schema does not define (mc:Ignorable lists them). Mutates in
    place. Re-implemented here (schemas vendored, code
    ours)."""
    for el in root.iter():
        for attr in list(el.attrib):
            ns = _ns_of(attr)
            if ns and ns not in _ALLOWED_OOXML_NS:
                del el.attrib[attr]

    def prune(node):
        for child in list(node):
            tag = child.tag
            if not isinstance(tag, str):  # comment / PI
                continue
            ns = _ns_of(tag)
            if ns and ns not in _ALLOWED_OOXML_NS:
                node.remove(child)
                continue
            prune(child)

    prune(root)


def validate_docx_xsd(path: str) -> None:
    """Validate a DOCX's `word/document.xml` against the bundled ISO/IEC 29500-4
    (Transitional) WML schema. No-op if the schemas are not vendored. Raises
    ValidationError on a schema violation (callers may choose to warn instead)."""
    global _wml_schema
    if not _WML_XSD.is_file():
        log.debug("OOXML schemas not vendored — skipping XSD validation")
        return
    from lxml import etree

    p = Path(path)
    try:
        with zipfile.ZipFile(p) as z:
            doc = etree.fromstring(z.read("word/document.xml"))
    except (zipfile.BadZipFile, KeyError, etree.XMLSyntaxError) as e:
        raise ValidationError(f"cannot read word/document.xml: {e}") from e

    _strip_ignorable(doc)

    if _wml_schema is None:
        _wml_schema = etree.XMLSchema(etree.parse(str(_WML_XSD)))
    if not _wml_schema.validate(doc):
        first = next(iter(_wml_schema.error_log), None)
        raise ValidationError(f"DOCX fails OOXML XSD: {first.message if first else 'invalid'}")


def validate_xlsx(path: str) -> None:
    p = Path(path)
    if not p.is_file() or p.stat().st_size == 0:
        raise ValidationError("XLSX is missing or empty")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(p))
    except Exception as e:
        raise ValidationError(f"XLSX failed to open: {e}") from e
    if not wb.sheetnames:
        raise ValidationError("XLSX has no sheets")


def validate_pdf(path: str) -> None:
    p = Path(path)
    if not p.is_file() or p.stat().st_size == 0:
        raise ValidationError("PDF is missing or empty")
    try:
        from pypdf import PdfReader

        pages = len(PdfReader(str(p)).pages)
    except Exception as e:
        raise ValidationError(f"PDF failed to open: {e}") from e
    if pages < 1:
        raise ValidationError("PDF has no pages")


# Attributes that cause the browser to LOAD a resource (the zero-egress risk).
# `<a href>` is excluded — a hyperlink does not fetch on render and is legitimate
# report content. Namespace URIs (xmlns) and URL-like strings inside inlined
# <script> bodies are ignored because we walk parsed elements, not raw text.
_LOAD_ATTRS = ("src", "srcset", "poster", "data-src")
_HTTP = re.compile(r"^\s*https?://", re.IGNORECASE)
_CSS_URL = re.compile(r"""(?:url\(\s*['"]?\s*|@import\s+['"])https?://""", re.IGNORECASE)


def validate_html(path: str) -> None:
    """A self-contained HTML artefact: parses, has NO external resource URL (the
    zero-egress / offline-portable guarantee), and every embedded JSON data island
    is valid. The strict CSP <meta> the html engine injects is the runtime enforcer;
    this is the deterministic catch for a model that wrote a CDN link anyway."""
    p = Path(path)
    if not p.is_file() or p.stat().st_size == 0:
        raise ValidationError("HTML is missing or empty")
    data = p.read_text(encoding="utf-8", errors="replace")
    try:
        from lxml import html as lxml_html

        doc = lxml_html.fromstring(data)
    except Exception as e:
        raise ValidationError(f"HTML failed to parse: {e}") from e

    offenders: list[str] = []
    for el in doc.iter():
        tag = el.tag if isinstance(el.tag, str) else ""
        # <link href> loads (stylesheet/preload/icon); <a href> does not.
        href = el.get("href")
        if href and tag != "a" and _HTTP.match(href):
            offenders.append(f"<{tag} href={href[:60]}>")
        for attr in _LOAD_ATTRS:
            v = el.get(attr)
            if v and _HTTP.match(v):
                offenders.append(f"<{tag} {attr}={v[:60]}>")
        style = el.get("style")
        if style and _CSS_URL.search(style):
            offenders.append(f"<{tag} style url()>")
    for st in doc.iter("style"):
        if st.text and _CSS_URL.search(st.text):
            offenders.append("<style url()/@import>")
    if offenders:
        raise ValidationError(f"external resource URL(s) (zero-egress): {offenders[:5]}")

    for sc in doc.iter("script"):
        if (sc.get("type") or "").strip().lower() == "application/json":
            try:
                json.loads(sc.text or "")
            except Exception as e:
                raise ValidationError(f"invalid JSON data island: {e}") from e
