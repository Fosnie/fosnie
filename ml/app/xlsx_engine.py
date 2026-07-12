# Copyright 2026 Private AI Ltd (SC881079)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Spreadsheet artefacts.
The model emits a **JSON workbook spec** (taught by the `xlsx-tables` skill); this
builds it into a real `.xlsx` via openpyxl — styled header row, per-column number
formats, and live formulas (a cell value beginning `=` becomes a formula, so it
recalculates in any spreadsheet app)."""

from __future__ import annotations

import json
from pathlib import Path

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _parse(content: str, title: str) -> dict:
    """Accept `{sheets:[…]}`, a single `{name?,columns?,rows}` sheet, or a bare list
    of rows. Raises ValueError (→ HTTP 400) on anything else."""
    try:
        spec = json.loads(content.strip())
    except Exception as e:
        raise ValueError(f"xlsx content must be a JSON workbook spec: {e}") from e

    if isinstance(spec, list):
        return {"sheets": [{"name": title or "Sheet", "rows": spec}]}
    if not isinstance(spec, dict):
        raise ValueError("xlsx spec must be a JSON object or array")
    if "sheets" in spec:
        if not isinstance(spec["sheets"], list) or not spec["sheets"]:
            raise ValueError("xlsx spec 'sheets' must be a non-empty array")
        return spec
    if "rows" in spec or "columns" in spec:
        return {"sheets": [{"name": spec.get("name", title or "Sheet"), **spec}]}
    raise ValueError("xlsx spec needs a 'sheets' array (or a 'rows'/'columns' sheet)")


def _header(col) -> str:
    return str(col.get("header", "")) if isinstance(col, dict) else str(col)


def build(content: str, title: str, out_path: str) -> dict:
    spec = _parse(content, title)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # start clean; we add named sheets below

    for sheet in spec["sheets"]:
        if not isinstance(sheet, dict):
            raise ValueError("each sheet must be an object")
        ws = wb.create_sheet(title=(str(sheet.get("name") or "Sheet"))[:31])
        cols = sheet.get("columns") or []

        if cols:
            ws.append([_header(c) for c in cols])
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFFFF")
                cell.fill = PatternFill("solid", fgColor="FF1F2933")
                cell.alignment = Alignment(vertical="center")

        for row in sheet.get("rows", []):
            if not isinstance(row, list):
                raise ValueError("each row must be an array of cell values")
            ws.append(row)  # a string beginning '=' is written as a live formula

        # Per-column number formats + a readable width.
        for i, col in enumerate(cols, start=1):
            letter = get_column_letter(i)
            if isinstance(col, dict) and col.get("format"):
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=i).number_format = str(col["format"])
            width = max([len(_header(col))] + [
                len(str(ws.cell(row=r, column=i).value or "")) for r in range(2, min(ws.max_row, 50) + 1)
            ] + [8])
            ws.column_dimensions[letter].width = min(width + 2, 48)

        if cols:
            ws.freeze_panes = "A2"  # keep the header visible when scrolling

    if not wb.sheetnames:
        wb.create_sheet("Sheet")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return {"path": str(out), "mime": XLSX_MIME}
